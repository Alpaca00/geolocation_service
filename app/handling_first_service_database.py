from pymongo import MongoClient
import openpyxl
from datetime import datetime, timedelta


client = MongoClient('localhost', 27017)
db = client.first_service
collection_rides = db.rides
collection_payments = db.payments



dir_rides = f"/home/oleg/python/geolocation_service/new_file_{datetime.now().strftime('%H_%M')}.xlsx"
dir_payments = f"/home/oleg/python/geolocation_service/new_file_{datetime.now().strftime('%H_%M')}.xlsx"


class FirstServiceDatabase:
    def __init__(self):
        self.data_rides = openpyxl.load_workbook(dir_rides)
        self.data_payments = openpyxl.load_workbook(dir_payments)


    def insert_document_rides_to_collection(self):
        self.trash_remove(collection_rides)
        wb = self.data_rides.active
        try:
            for row in wb.iter_rows(values_only=True):
                collection_rides.insert_one({
                        'driver_id': row[0], 'license_plato': row[1], 'distance': row[2],
                        'cost': row[3], 'cash': row[4], 'cashless': row[5],
                        'pickup_time': row[6], 'route': row[7], 'status': row[8],
                        'type_payment': row[9], 'profit': row[10]
                    })
        except IndexError:
            return True

    def insert_document_payments_to_collection(self):
        self.trash_remove(collection_payments)
        wb = self.data_payments.active
        try:
            for row in wb.iter_rows(values_only=True):
                collection_payments.insert_one({
                    'driver_id': row[0], 'license_plato': row[1], 'total_cashless': row[2],
                    'total_fee': row[5], 'profit': row[6], 'total_bonus': row[7],
                    'total_cash': row[3], 'total_cost': row[4], 'mileage': row[8]
                })
        except IndexError:
            return True

    @staticmethod
    def trash_remove(c):
        remove = c.delete_many({})
        print(remove.deleted_count, " documents deleted.")

    @staticmethod
    def checking_collections():
        query_commission = collection_payments.find()
        commission_all = 0
        for i in query_commission:
            if isinstance(i['total_fee'], float):
                commission_all += i['total_fee']

        query_total_cost = collection_rides.find()
        arr = []
        for item in query_total_cost:
            if isinstance(item['cost'], int):
                arr.append(item['cost'])
        total_cost = 0
        for pcs in arr:
            total_cost += pcs

        query_profit = collection_payments.find()
        profit = 0
        for pr in query_profit:
            if isinstance(pr['profit'], float):
                profit += pr['profit']

        result = total_cost - commission_all
        return result == profit



